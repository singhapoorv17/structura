"""Presentation conventions for the Structura workbook.

The colour convention is the one used on every project-finance desk, and a
credit officer reads it without being told:

===========  ==========================================================
Blue font    **Input.** Change it. Nothing upstream feeds it.
Black font   **Formula.** Calculated on this sheet from cells on this sheet.
Green font   **Link.** A formula whose only job is to pull a value in from
             another sheet. Keeping links visually distinct is what makes a
             model auditable: you can see at a glance where a sheet's
             boundary is.
Amber fill   **Solver-derived input.** A blue input cell whose value came out
             of Structura's Python solver rather than out of the sheet,
             because the quantity cannot be expressed as an Excel formula.
             Every one of these is listed on the Notes sheet.
===========  ==========================================================

Number formats follow banker convention too: whole dollars with negatives in
brackets, coverage ratios as ``1.30x``, rates to one or three decimals as
appropriate. Nothing is shown to a precision the model does not have.
"""

from __future__ import annotations

from typing import Final

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

__all__ = [
    "FONT_NAME",
    "MONEY",
    "MONEY_M",
    "RATIO",
    "PERCENT_1",
    "PERCENT_2",
    "RATE_3",
    "NUMBER_0",
    "MONEY_OR_NO_LIMIT",
    "NUMBER_2",
    "PRICE",
    "DATE_FMT",
    "YEARS",
    "TITLE_FONT",
    "SUBTITLE_FONT",
    "SECTION_FONT",
    "HEADER_FONT",
    "LABEL_FONT",
    "INPUT_FONT",
    "FORMULA_FONT",
    "LINK_FONT",
    "TOTAL_FONT",
    "NOTE_FONT",
    "HEADER_FILL",
    "SECTION_FILL",
    "INPUT_FILL",
    "SOLVER_FILL",
    "CHECK_FILL",
    "TOTAL_BORDER",
    "SECTION_BORDER",
    "BOX_BORDER",
    "LEFT",
    "RIGHT",
    "CENTRE",
    "CellStyle",
    "TITLE",
    "SUBTITLE",
    "SECTION",
    "HEADER",
    "LABEL",
    "SUBHEAD",
    "UNIT",
    "INPUT",
    "SOLVER_INPUT",
    "FORMULA",
    "LINK",
    "TOTAL",
    "NOTE",
    "CHECK",
    "COL_LABEL",
    "COL_UNIT",
    "COL_TOTAL",
    "COL_FIRST_PERIOD",
    "LABEL_WIDTH",
    "UNIT_WIDTH",
    "TOTAL_WIDTH",
    "PERIOD_WIDTH",
]

FONT_NAME: Final[str] = "Calibri"

# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------

#: Whole dollars, thousands separated, negatives in brackets. Project-finance
#: numbers are eight and nine figures; cents are noise and printing them
#: implies a precision the model does not have.
MONEY: Final[str] = "#,##0;(#,##0)"
#: Dollars in millions, one decimal - the unit a credit committee talks in.
MONEY_M: Final[str] = '#,##0.0,,"m";(#,##0.0,,"m")'
#: Coverage ratios. 1.30x, not 1.3 or 130%.
RATIO: Final[str] = '0.000"x"'
PERCENT_1: Final[str] = "0.0%"
PERCENT_2: Final[str] = "0.00%"
#: Interest rates and spreads, where a basis point matters.
RATE_3: Final[str] = "0.000%"
NUMBER_0: Final[str] = "#,##0"
#: Money, except that the "no numeric limit applies" sentinel prints as words.
#: Excel has no infinity literal, so a credit test that does not bind is
#: carried as an unreachable magnitude; printing 1E+30 in a dollar column would
#: look like a bug rather than like "not tested".
MONEY_OR_NO_LIMIT: Final[str] = '[>1000000000000]"no limit";#,##0'

NUMBER_2: Final[str] = "#,##0.00"
PRICE: Final[str] = '"$"#,##0.00'
DATE_FMT: Final[str] = "dd-mmm-yy"
YEARS: Final[str] = '0.0" y"'

# ---------------------------------------------------------------------------
# Fonts, fills and borders
# ---------------------------------------------------------------------------

_NAVY: Final[str] = "FF1F3864"
_BLUE: Final[str] = "FF0000CC"
_GREEN: Final[str] = "FF006600"
_BLACK: Final[str] = "FF000000"
_GREY: Final[str] = "FF7F7F7F"

TITLE_FONT: Final = Font(name=FONT_NAME, size=14, bold=True, color=_NAVY)
SUBTITLE_FONT: Final = Font(name=FONT_NAME, size=9, italic=True, color=_GREY)
SECTION_FONT: Final = Font(name=FONT_NAME, size=10, bold=True, color=_NAVY)
HEADER_FONT: Final = Font(name=FONT_NAME, size=9, bold=True, color="FFFFFFFF")
LABEL_FONT: Final = Font(name=FONT_NAME, size=10, color=_BLACK)
INPUT_FONT: Final = Font(name=FONT_NAME, size=10, color=_BLUE)
FORMULA_FONT: Final = Font(name=FONT_NAME, size=10, color=_BLACK)
LINK_FONT: Final = Font(name=FONT_NAME, size=10, color=_GREEN)
TOTAL_FONT: Final = Font(name=FONT_NAME, size=10, bold=True, color=_BLACK)
NOTE_FONT: Final = Font(name=FONT_NAME, size=9, italic=True, color=_GREY)

HEADER_FILL: Final = PatternFill("solid", fgColor=_NAVY)
SECTION_FILL: Final = PatternFill("solid", fgColor="FFDCE6F1")
INPUT_FILL: Final = PatternFill("solid", fgColor="FFFFFFCC")
#: Amber: an input cell whose value came from the Python solver, not the sheet.
SOLVER_FILL: Final = PatternFill("solid", fgColor="FFFCE4D6")
CHECK_FILL: Final = PatternFill("solid", fgColor="FFEAF3EA")

_THIN: Final = Side(style="thin", color=_NAVY)
_DOUBLE: Final = Side(style="double", color=_NAVY)
_HAIR: Final = Side(style="hair", color=_GREY)

TOTAL_BORDER: Final = Border(top=_THIN, bottom=_DOUBLE)
SECTION_BORDER: Final = Border(bottom=_THIN)
BOX_BORDER: Final = Border(top=_HAIR, bottom=_HAIR, left=_HAIR, right=_HAIR)

LEFT: Final = Alignment(horizontal="left", vertical="center")
RIGHT: Final = Alignment(horizontal="right", vertical="center")
CENTRE: Final = Alignment(horizontal="center", vertical="center")


class CellStyle:
    """A bundle of font / fill / border / alignment applied to one cell.

    Deliberately not an openpyxl ``NamedStyle``: named styles are workbook
    singletons and cannot vary their number format per cell, which every
    financial model needs.
    """

    __slots__ = ("font", "fill", "border", "alignment", "number_format")

    def __init__(
        self,
        font: Font,
        *,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        number_format: str | None = None,
    ) -> None:
        self.font = font
        self.fill = fill
        self.border = border
        self.alignment = alignment
        self.number_format = number_format

    def apply(self, cell, number_format: str | None = None) -> None:
        """Stamp this style onto ``cell``, with an optional format override."""
        cell.font = self.font
        if self.fill is not None:
            cell.fill = self.fill
        if self.border is not None:
            cell.border = self.border
        if self.alignment is not None:
            cell.alignment = self.alignment
        fmt = number_format or self.number_format
        if fmt is not None:
            cell.number_format = fmt


TITLE = CellStyle(TITLE_FONT, alignment=LEFT)
SUBTITLE = CellStyle(SUBTITLE_FONT, alignment=LEFT)
SECTION = CellStyle(
    SECTION_FONT, fill=SECTION_FILL, border=SECTION_BORDER, alignment=LEFT
)
HEADER = CellStyle(HEADER_FONT, fill=HEADER_FILL, alignment=CENTRE)
LABEL = CellStyle(LABEL_FONT, alignment=LEFT)
SUBHEAD = CellStyle(TOTAL_FONT, alignment=LEFT)
UNIT = CellStyle(NOTE_FONT, alignment=LEFT)
INPUT = CellStyle(INPUT_FONT, fill=INPUT_FILL, border=BOX_BORDER, alignment=RIGHT)
SOLVER_INPUT = CellStyle(
    INPUT_FONT, fill=SOLVER_FILL, border=BOX_BORDER, alignment=RIGHT
)
FORMULA = CellStyle(FORMULA_FONT, alignment=RIGHT)
LINK = CellStyle(LINK_FONT, alignment=RIGHT)
TOTAL = CellStyle(TOTAL_FONT, border=TOTAL_BORDER, alignment=RIGHT)
NOTE = CellStyle(NOTE_FONT, alignment=LEFT)
CHECK = CellStyle(FORMULA_FONT, fill=CHECK_FILL, alignment=RIGHT)

# ---------------------------------------------------------------------------
# Column geometry, shared by every time-series sheet
# ---------------------------------------------------------------------------

#: A = line-item label, B = unit, C = total (or the value of a scalar row),
#: D onwards = one column per model period. Every sheet in the workbook uses
#: the same geometry so a reader's eye does not have to re-anchor.
COL_LABEL: Final[int] = 1
COL_UNIT: Final[int] = 2
COL_TOTAL: Final[int] = 3
COL_FIRST_PERIOD: Final[int] = 4

LABEL_WIDTH: Final[float] = 46.0
UNIT_WIDTH: Final[float] = 11.0
TOTAL_WIDTH: Final[float] = 16.0
PERIOD_WIDTH: Final[float] = 13.5
