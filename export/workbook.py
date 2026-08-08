"""The workbook builder: iterative calculation, cell helpers and a sheet registry.

WHY THIS FILE EXISTS AT ALL
---------------------------
``openpyxl`` is used for one reason: it is the only writer that emits
``iterate="1"`` into ``xl/workbook.xml``. A lender-grade project
finance model has genuine circular references - IDC depends on the debt draw,
the draw depends on total project cost, total project cost includes IDC - and
without iterative calculation Excel opens the file with a circular-reference
warning and zeroes the chain. The first thing this builder does is therefore
set ``CalcProperties(iterate=True, ...)``.

``fullCalcOnLoad=True`` is equally mandatory but for a different reason:
openpyxl writes ``<f>`` elements with **no cached ``<v>``**. Every formula cell
in a file it produces is literally valueless until Excel calculates it. Without
``fullCalcOnLoad`` the workbook opens showing zeros.

THE REFERENCE REGISTRY
----------------------
Sheets have to point at each other's rows: the Debt sheet divides the
Operations sheet's CFADS row, the Waterfall sheet reads the Debt sheet's
principal row. Hard-coding ``'Operations'!$D$27`` in another module makes the
workbook unmaintainable - insert one row and everything silently breaks.

So every row a sheet writes is registered under a stable logical key
(``"operations.cfads"``), and every other sheet asks the builder for the
address. Inserting a row anywhere changes nothing downstream.

Scalars use Excel **defined names** instead (``Senior_Debt``,
``Total_Project_Cost``). A named range is self-documenting inside the formula
bar, which is exactly what an auditor tracing a number wants to see.

THE SHEET REGISTRY
------------------
Phases 2 and 3 add a Tax sheet and a Structure-comparison sheet. They must be
able to do that without touching this file or any existing sheet module, so
sheets register themselves::

    register_sheet("Tax", build_tax, display_order=60, build_order=45)

``display_order`` is where the tab sits; ``build_order`` is when the builder
runs, which matters only because a sheet must be built after any sheet whose
*rows* (not names) it references.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Callable, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet

from export import styles
from export.styles import (
    COL_FIRST_PERIOD,
    COL_LABEL,
    COL_TOTAL,
    COL_UNIT,
    CellStyle,
)

__all__ = [
    "CALC_ID",
    "ITERATE_COUNT",
    "ITERATE_DELTA",
    "INFINITY_SENTINEL",
    "RowRef",
    "SheetSpec",
    "SheetWriter",
    "WorkbookBuilder",
    "register_sheet",
    "registered_sheets",
    "clear_registry",
]

#: Excel's own ``calcId`` for the 2019+ calculation engine. Writing a modern id
#: stops Excel treating the file as produced by a legacy calculation chain.
CALC_ID: int = 191029
#: Iteration budget for the circular funding chain. The IDC / fee / DSRA loop
#: has a gain of roughly ``rate x construction years`` - a few percent - so it
#: converges geometrically in well under twenty passes. 100 is generous.
ITERATE_COUNT: int = 100
#: Convergence threshold in currency units. Project cashflows are 1e8-1e9, so
#: a 1e-4 absolute delta is ~1e-13 relative: exact for every practical purpose.
ITERATE_DELTA: float = 0.0001

#: Stand-in for "this test imposes no limit". Excel has no infinity literal and
#: ``MIN()`` over a range must stay numeric, so an unreachable magnitude is
#: used instead. Any real facility is many orders of magnitude below it.
INFINITY_SENTINEL: str = "1E+30"


# ---------------------------------------------------------------------------
# Row references
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class RowRef:
    """Where a registered line item physically lives.

    ``start_period`` is 1 on every time-series sheet except Returns, whose
    equity cashflow series has to start at period 0 (the COD investment).
    """

    sheet: str
    row: int
    first_col: int
    n_periods: int
    start_period: int = 1

    def column(self, period: int) -> int:
        """1-based worksheet column holding ``period``."""
        col = self.first_col + (period - self.start_period)
        if not self.first_col <= col < self.first_col + self.n_periods:
            raise IndexError(
                f"period {period} is outside {self.sheet!r} row {self.row} "
                f"(periods {self.start_period}"
                f"..{self.start_period + self.n_periods - 1})"
            )
        return col

    def has(self, period: int) -> bool:
        """True if ``period`` has a column on this row."""
        return (
            self.start_period <= period <= self.start_period + self.n_periods - 1
        )

    def cell(self, period: int, absolute: bool = True) -> str:
        """Fully qualified reference to one period, e.g. ``'Debt'!$G$31``."""
        return _address(self.sheet, self.row, self.column(period), absolute)

    def span(
        self, first: int | None = None, last: int | None = None, absolute: bool = True
    ) -> str:
        """Fully qualified reference to a contiguous run of periods."""
        first = self.start_period if first is None else first
        last = (
            self.start_period + self.n_periods - 1 if last is None else last
        )
        if last < first:
            raise ValueError(f"empty span {first}..{last} on {self.sheet!r}")
        return (
            f"{_address(self.sheet, self.row, self.column(first), absolute)}"
            f":{_local_address(self.row, self.column(last), absolute)}"
        )


def _local_address(row: int, col: int, absolute: bool = True) -> str:
    marker = "$" if absolute else ""
    return f"{marker}{get_column_letter(col)}{marker}{row}"


def _address(sheet: str, row: int, col: int, absolute: bool = True) -> str:
    return f"'{sheet}'!{_local_address(row, col, absolute)}"


# ---------------------------------------------------------------------------
# Sheet registry
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class SheetSpec:
    """One sheet's identity and the callable that fills it."""

    name: str
    builder: Callable[["WorkbookBuilder", Any], None]
    display_order: int
    build_order: int
    description: str = ""


_REGISTRY: dict[str, SheetSpec] = {}


def register_sheet(
    name: str,
    builder: Callable[["WorkbookBuilder", Any], None],
    *,
    display_order: int,
    build_order: int | None = None,
    description: str = "",
    replace: bool = False,
) -> SheetSpec:
    """Register a sheet so :func:`export.api.build_workbook` will emit it.

    Parameters
    ----------
    name:
        Worksheet tab name, and the name used in cross-sheet references.
    builder:
        ``builder(wb: WorkbookBuilder, model) -> None``. It receives the shared
        builder and the model bundle, and writes into ``wb.sheet(name)``.
    display_order:
        Tab position. Existing sheets are spaced ten apart so a later phase can
        slot in without renumbering (Tax is intended at 55, Structures at 75).
    build_order:
        When the builder runs. Defaults to ``display_order``. Raise it above a
        sheet whose *rows* this sheet references; defined names need no
        ordering because they are resolved by Excel, not by this module.
    replace:
        Guard against two phases silently claiming the same tab name.
    """
    if name in _REGISTRY and not replace:
        raise ValueError(
            f"sheet {name!r} is already registered; pass replace=True to override"
        )
    spec = SheetSpec(
        name=name,
        builder=builder,
        display_order=display_order,
        build_order=display_order if build_order is None else build_order,
        description=description,
    )
    _REGISTRY[name] = spec
    return spec


def registered_sheets() -> tuple[SheetSpec, ...]:
    """Every registered sheet, in display order."""
    return tuple(sorted(_REGISTRY.values(), key=lambda s: (s.display_order, s.name)))


def clear_registry() -> None:
    """Drop every registration. Test-support only."""
    _REGISTRY.clear()


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------


class SheetWriter:
    """Row-oriented writer for one worksheet.

    Holds a cursor so sheet modules never compute row numbers by hand, and
    registers every row it writes with the parent builder.
    """

    def __init__(
        self,
        builder: "WorkbookBuilder",
        ws: Worksheet,
        *,
        n_periods: int,
        start_period: int = 1,
    ) -> None:
        self.builder = builder
        self.ws = ws
        self.n_periods = n_periods
        self.start_period = start_period
        self.row = 1

    # -- geometry -----------------------------------------------------------

    @property
    def name(self) -> str:
        return self.ws.title

    @property
    def last_period(self) -> int:
        return self.start_period + self.n_periods - 1

    def periods(self) -> range:
        return range(self.start_period, self.last_period + 1)

    def column(self, period: int) -> int:
        return COL_FIRST_PERIOD + (period - self.start_period)

    def letter(self, period: int) -> str:
        return get_column_letter(self.column(period))

    def local(self, row: int, period: int, absolute: bool = False) -> str:
        """Same-sheet reference to ``(row, period)`` - used inside formulas."""
        return _local_address(row, self.column(period), absolute)

    # -- layout -------------------------------------------------------------

    def set_widths(
        self,
        *,
        label: float = styles.LABEL_WIDTH,
        unit: float = styles.UNIT_WIDTH,
        total: float = styles.TOTAL_WIDTH,
        period: float = styles.PERIOD_WIDTH,
    ) -> None:
        dims = self.ws.column_dimensions
        dims[get_column_letter(COL_LABEL)].width = label
        dims[get_column_letter(COL_UNIT)].width = unit
        dims[get_column_letter(COL_TOTAL)].width = total
        for p in self.periods():
            dims[self.letter(p)].width = period

    def hide_gridlines(self) -> None:
        """Gridlines off. A model that shows them looks like a scratchpad."""
        self.ws.sheet_view.showGridLines = False

    def freeze(self, cell: str) -> None:
        self.ws.freeze_panes = cell

    def skip(self, rows: int = 1) -> None:
        self.row += rows

    # -- content ------------------------------------------------------------

    def title_block(self, title: str, subtitle: str = "") -> None:
        self.write(COL_LABEL, self.row, title, styles.TITLE)
        self.row += 1
        if subtitle:
            self.write(COL_LABEL, self.row, subtitle, styles.SUBTITLE)
            self.row += 1
        self.row += 1

    def section(self, label: str, width: int | None = None) -> int:
        """A shaded, bold, underlined section header spanning the used width."""
        row = self.row
        last = (
            COL_FIRST_PERIOD + self.n_periods - 1 if width is None else width
        )
        self.write(COL_LABEL, row, label, styles.SECTION)
        for col in range(COL_UNIT, last + 1):
            styles.SECTION.apply(self.ws.cell(row=row, column=col))
        self.row += 1
        return row

    def period_header(
        self,
        *,
        label: str = "Model period",
        total_label: str = "Total",
        dates: Sequence[date] | None = None,
    ) -> None:
        """The banded period header, optionally with a period-end date strip."""
        row = self.row
        self.write(COL_LABEL, row, label, styles.HEADER)
        self.write(COL_UNIT, row, "", styles.HEADER)
        self.write(COL_TOTAL, row, total_label, styles.HEADER)
        for p in self.periods():
            self.write(self.column(p), row, p, styles.HEADER, styles.NUMBER_0)
        self.row += 1
        if dates is not None:
            drow = self.row
            self.write(COL_LABEL, drow, "Period ending", styles.NOTE)
            for i, p in enumerate(self.periods()):
                if i < len(dates):
                    self.write(
                        self.column(p),
                        drow,
                        dates[i],
                        styles.NOTE,
                        styles.DATE_FMT,
                    )
            self.row += 1

    def write(
        self,
        col: int,
        row: int,
        value: Any,
        style: CellStyle,
        number_format: str | None = None,
    ):
        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        style.apply(cell, number_format)
        return cell

    # -- scalar rows --------------------------------------------------------

    def scalar(
        self,
        label: str,
        value: Any,
        *,
        unit: str = "",
        style: CellStyle = styles.FORMULA,
        number_format: str | None = None,
        name: str | None = None,
        key: str | None = None,
        label_style: CellStyle = styles.LABEL,
        source: str = "",
    ) -> int:
        """One labelled value in the total column. Returns the row number.

        ``name`` defines an Excel named range over the value cell, which is how
        every other sheet should refer to it. ``source`` is free text written to
        the right of the value - the citation or the note that makes the number
        auditable.
        """
        row = self.row
        self.write(COL_LABEL, row, label, label_style)
        if unit:
            self.write(COL_UNIT, row, unit, styles.UNIT)
        self.write(COL_TOTAL, row, value, style, number_format)
        if source:
            self.write(COL_FIRST_PERIOD, row, source, styles.NOTE)
        if name:
            self.builder.define_name(name, self.name, row, COL_TOTAL)
        if key:
            self.builder.register_cell(key, self.name, row, COL_TOTAL)
        self.row += 1
        return row

    def note(self, text: str) -> int:
        row = self.row
        self.write(COL_LABEL, row, text, styles.NOTE)
        self.row += 1
        return row

    # -- period rows --------------------------------------------------------

    def values_row(
        self,
        key: str | None,
        label: str,
        values: Iterable[Any],
        *,
        unit: str = "",
        style: CellStyle = styles.INPUT,
        number_format: str | None = None,
        total: bool = False,
        series_name: str | None = None,
    ) -> RowRef:
        """A row of literal values - i.e. inputs, or an index strip."""
        row = self.row
        self.write(COL_LABEL, row, label, styles.LABEL)
        if unit:
            self.write(COL_UNIT, row, unit, styles.UNIT)
        seq = list(values)
        for i, p in enumerate(self.periods()):
            value = seq[i] if i < len(seq) else None
            self.write(self.column(p), row, value, style, number_format)
        ref = RowRef(
            self.name, row, COL_FIRST_PERIOD, self.n_periods, self.start_period
        )
        if total:
            self._write_total(row, ref, number_format)
        self.row += 1
        if key:
            self.builder.register_row(key, ref)
        if series_name:
            self._define_series(series_name, ref)
        return ref

    def formula_row(
        self,
        key: str | None,
        label: str,
        formula: Callable[[int], str | None],
        *,
        unit: str = "",
        style: CellStyle = styles.FORMULA,
        number_format: str | None = None,
        total: bool = False,
        series_name: str | None = None,
    ) -> RowRef:
        """A row of formulas. ``formula(period)`` returns the text after ``=``.

        Returning ``None`` leaves the cell empty, which is what a row wants at
        the far edge of a lookahead window.
        """
        row = self.row
        self.write(COL_LABEL, row, label, styles.LABEL)
        if unit:
            self.write(COL_UNIT, row, unit, styles.UNIT)
        for p in self.periods():
            body = formula(p)
            self.write(
                self.column(p),
                row,
                None if body is None else f"={body}",
                style,
                number_format,
            )
        ref = RowRef(
            self.name, row, COL_FIRST_PERIOD, self.n_periods, self.start_period
        )
        if total:
            self._write_total(row, ref, number_format)
        self.row += 1
        if key:
            self.builder.register_row(key, ref)
        if series_name:
            self._define_series(series_name, ref)
        return ref

    def _define_series(self, name: str, ref: RowRef) -> None:
        """Name the whole period span of a row.

        A reader who selects ``CFADS_Series`` in the name box lands on the
        CFADS row of the Operations sheet. It also gives downstream formulas a
        readable way to index into a series with ``INDEX()``.
        """
        self.builder.define_range(
            name,
            self.name,
            ref.row,
            ref.first_col,
            ref.first_col + ref.n_periods - 1,
        )

    def _write_total(
        self, row: int, ref: RowRef, number_format: str | None
    ) -> None:
        local = (
            f"{_local_address(row, ref.first_col)}:"
            f"{_local_address(row, ref.first_col + ref.n_periods - 1)}"
        )
        self.write(
            COL_TOTAL, row, f"=SUM({local})", styles.TOTAL, number_format
        )


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------


@dataclass
class WorkbookBuilder:
    """Owns the workbook, the calculation properties and the reference maps."""

    wb: Workbook = field(default_factory=Workbook)
    _rows: dict[str, RowRef] = field(default_factory=dict)
    _cells: dict[str, str] = field(default_factory=dict)
    _names: dict[str, str] = field(default_factory=dict)
    _sheets: dict[str, SheetWriter] = field(default_factory=dict)

    def __post_init__(self) -> None:
        # openpyxl seeds a workbook with one blank sheet; every sheet here is
        # created explicitly, so drop it.
        for ws in list(self.wb.worksheets):
            self.wb.remove(ws)
        self.enable_iterative_calculation()

    # -- calculation properties --------------------------------------------

    def enable_iterative_calculation(self) -> None:
        """Switch on iterative calculation and force a full calc on open.

        Both halves are required:

        * ``iterate`` lets Excel resolve the IDC <-> debt size <-> fees <->
          DSRA circularity natively, instead of raising a circular-reference
          warning and zeroing the chain.
        * ``fullCalcOnLoad`` is required because openpyxl writes formulas with
          no cached value. Without it the workbook opens showing zeros.
        """
        self.wb.calculation = CalcProperties(
            calcId=CALC_ID,
            fullCalcOnLoad=True,
            iterate=True,
            iterateCount=ITERATE_COUNT,
            iterateDelta=ITERATE_DELTA,
        )

    # -- sheets -------------------------------------------------------------

    def create_sheet(
        self, name: str, *, n_periods: int, start_period: int = 1
    ) -> SheetWriter:
        ws = self.wb.create_sheet(title=name)
        writer = SheetWriter(
            self, ws, n_periods=n_periods, start_period=start_period
        )
        self._sheets[name] = writer
        return writer

    def sheet(self, name: str) -> SheetWriter:
        return self._sheets[name]

    def has_sheet(self, name: str) -> bool:
        return name in self._sheets

    # -- reference registry -------------------------------------------------

    def register_row(self, key: str, ref: RowRef) -> None:
        if key in self._rows:
            raise ValueError(f"row key {key!r} is already registered")
        self._rows[key] = ref

    def row(self, key: str) -> RowRef:
        try:
            return self._rows[key]
        except KeyError as exc:
            raise KeyError(
                f"unknown row key {key!r}; is the sheet that writes it built "
                f"earlier in the build order?"
            ) from exc

    def ref(self, key: str, period: int) -> str:
        """Absolute reference to one period of a registered row."""
        return self.row(key).cell(period)

    def span(self, key: str, first: int | None = None, last: int | None = None) -> str:
        """Absolute reference to a contiguous run of a registered row."""
        return self.row(key).span(first, last)

    def register_cell(self, key: str, sheet: str, row: int, col: int) -> None:
        self._cells[key] = _address(sheet, row, col)

    def cell(self, key: str) -> str:
        return self._cells[key]

    # -- defined names ------------------------------------------------------

    def define_name(self, name: str, sheet: str, row: int, col: int) -> str:
        """Create a workbook-scoped named range over a single cell.

        Named ranges are what make the formulas readable: ``=CFADS/Target_DSCR``
        instead of ``=D27/$C$18``. Every driver carries one.
        """
        if name in self._names:
            raise ValueError(f"defined name {name!r} already exists")
        _validate_name(name)
        ref = _address(sheet, row, col)
        self.wb.defined_names[name] = DefinedName(name, attr_text=ref)
        self._names[name] = ref
        return ref

    def define_range(
        self, name: str, sheet: str, row: int, first_col: int, last_col: int
    ) -> str:
        """Named range over a horizontal run of cells."""
        if name in self._names:
            raise ValueError(f"defined name {name!r} already exists")
        _validate_name(name)
        ref = (
            f"{_address(sheet, row, first_col)}:"
            f"{_local_address(row, last_col)}"
        )
        self.wb.defined_names[name] = DefinedName(name, attr_text=ref)
        self._names[name] = ref
        return ref

    def name_ref(self, name: str) -> str:
        return self._names[name]

    @property
    def names(self) -> dict[str, str]:
        return dict(self._names)

    # -- output -------------------------------------------------------------

    def order_sheets(self, order: Sequence[str]) -> None:
        """Arrange tabs. Sheets not named keep their relative position at the end."""
        index = {name: i for i, name in enumerate(order)}
        self.wb._sheets.sort(  # noqa: SLF001 - the only supported reorder hook
            key=lambda ws: index.get(ws.title, len(index))
        )

    def save(self, path) -> None:
        self.wb.save(path)


_RESERVED_NAME_PREFIXES = ("C", "R", "c", "r")


def _validate_name(name: str) -> None:
    """Reject names Excel would refuse or silently reinterpret."""
    if not name:
        raise ValueError("defined name cannot be empty")
    if not (name[0].isalpha() or name[0] == "_"):
        raise ValueError(f"defined name {name!r} must start with a letter or _")
    if any(not (ch.isalnum() or ch in "_.") for ch in name):
        raise ValueError(f"defined name {name!r} may only contain A-Z 0-9 _ .")
    if len(name) > 255:
        raise ValueError(f"defined name {name!r} exceeds 255 characters")
    if len(name) in (1, 2) and name in _RESERVED_NAME_PREFIXES:
        raise ValueError(f"{name!r} is reserved by Excel (R1C1 notation)")
    # A name that looks like a cell address (e.g. "LOG10" is fine, "AB12" is not)
    head = "".join(ch for ch in name if ch.isalpha())
    tail = name[len(head):]
    if head and tail.isdigit() and len(head) <= 3 and "_" not in name:
        raise ValueError(f"defined name {name!r} collides with a cell address")


def to_excel_date(value: date | datetime) -> datetime:
    """Excel stores dates as datetimes; normalise so openpyxl writes a serial."""
    if isinstance(value, datetime):
        return value
    return datetime(value.year, value.month, value.day)
