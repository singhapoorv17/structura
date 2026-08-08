"""Structural tests for the exported workbook.

The gate SPEC.md §9 makes blocking, the "live formulas, not pasted values"
requirement from §4.2 and §6.5, and the named-range requirement from §6.5.
"""

from __future__ import annotations

import re
import zipfile

import pytest
from openpyxl import load_workbook

from engine import AmortizationStyle, DebtTerms, ProjectInputs, TaxTreatment
from export import registered_sheets
from export.workbook import ITERATE_COUNT, ITERATE_DELTA
from test_export_evaluator import build_case, evaluated_case

# ---------------------------------------------------------------------------
# SPEC §9 gate: iterative calculation must reach xl/workbook.xml
# ---------------------------------------------------------------------------


def _calc_pr(path) -> str:
    xml = zipfile.ZipFile(path).read("xl/workbook.xml").decode()
    match = re.search(r"<calcPr[^>]*/?>", xml)
    assert match is not None, "no <calcPr> element in xl/workbook.xml"
    return match.group(0)


def test_workbook_xml_declares_iterative_calculation():
    """The blocking finding of SPEC.md §5 and §9.

    Without ``iterate="1"`` the workbook opens with a circular-reference
    warning and zeroes the IDC / debt-size / fee chain, and the entire
    lender-grade claim fails. openpyxl is the only writer that emits it.
    """
    path, _ = build_case("base")
    calc = _calc_pr(path)
    assert 'iterate="1"' in calc, calc
    assert f'iterateCount="{ITERATE_COUNT}"' in calc, calc
    assert 'iterateCount="100"' in calc, calc
    assert 'iterateDelta="0.0001"' in calc, calc
    assert str(ITERATE_DELTA) == "0.0001"


def test_workbook_xml_forces_a_full_calculation_on_load():
    """Mandatory: openpyxl writes ``<f>`` with no cached ``<v>``.

    Every formula cell in the file is valueless until Excel calculates it, so
    without ``fullCalcOnLoad`` the workbook opens showing zeros.
    """
    path, _ = build_case("base")
    assert 'fullCalcOnLoad="1"' in _calc_pr(path)


def test_no_formula_cell_carries_a_cached_value():
    """The corollary: if openpyxl ever started caching values, the
    ``fullCalcOnLoad`` requirement would need revisiting. Assert the premise."""
    path, _ = build_case("base")
    sheet_xml = [
        name
        for name in zipfile.ZipFile(path).namelist()
        if name.startswith("xl/worksheets/sheet")
    ]
    assert sheet_xml
    for name in sheet_xml:
        xml = zipfile.ZipFile(path).read(name).decode()
        assert "</f><v>" not in xml, f"{name} carries cached formula results"


# ---------------------------------------------------------------------------
# Sheets and registry
# ---------------------------------------------------------------------------


def test_expected_sheets_in_expected_order():
    path, _ = build_case("base")
    wb = load_workbook(path)
    assert wb.sheetnames == [
        "Summary",
        "Inputs",
        "Construction",
        "Operations",
        "Debt",
        "Waterfall",
        "Returns",
        "Notes",
    ]


def test_registry_leaves_room_for_the_deferred_phase_2_and_3_sheets():
    """Tax (Phase 2) and Structures (Phase 3) slot in without renumbering."""
    orders = {spec.name: spec.display_order for spec in registered_sheets()}
    assert orders["Debt"] < 55 < orders["Waterfall"], "no gap for a Tax sheet"
    assert orders["Waterfall"] < 75 < orders["Returns"], "no gap for Structures"


def test_a_new_sheet_can_be_registered_and_appears_in_the_output(tmp_path):
    """The extension point Phases 2 and 3 depend on."""
    from engine import run_model
    from export import build_workbook, register_sheet
    from export.sheets import register_default_sheets

    def build_demo(wb, model):
        sw = wb.create_sheet("Demo", n_periods=model.n_periods)
        sw.scalar("Senior debt, echoed", "=Senior_Debt", name="Demo_Echo")

    register_sheet("Demo", build_demo, display_order=55, build_order=65)
    try:
        project, terms = ProjectInputs(), DebtTerms()
        path = build_workbook(
            project, terms, run_model(project, terms), tmp_path / "demo.xlsx"
        )
        wb = load_workbook(path)
        assert wb.sheetnames.index("Demo") == wb.sheetnames.index("Debt") + 1
        assert "Demo_Echo" in wb.defined_names
    finally:
        from export.workbook import _REGISTRY

        _REGISTRY.pop("Demo", None)
        register_default_sheets()


# ---------------------------------------------------------------------------
# Live formulas, not pasted values
# ---------------------------------------------------------------------------

#: The rows whose cells must be formulas. If any of these were ever written as
#: numbers, changing an input in Excel would not move them, and the workbook
#: would be a report rather than a model.
_MUST_BE_FORMULAS = [
    ("Debt", "debt.opening_balance"),
    ("Debt", "debt.interest"),
    ("Debt", "debt.debt_service"),
    ("Debt", "debt.principal"),
    ("Debt", "debt.closing_balance"),
    ("Debt", "debt.dscr"),
    ("Debt", "debt.dsra_target"),
    ("Debt", "debt.available_service"),
    ("Operations", "ops.cfads"),
    ("Operations", "ops.revenue"),
    ("Operations", "ops.ebitda"),
    ("Construction", "con.interest"),
    ("Construction", "con.debt_draw"),
    ("Waterfall", "wf.distributions"),
    ("Waterfall", "wf.interest_paid"),
    ("Waterfall", "wf.principal_paid"),
    ("Returns", "ret.equity_cashflow"),
]


def _row_addresses(path):
    """Map every registered row key to its (sheet, row) - rebuilt from a live
    build so the test never hard-codes a row number."""
    from datetime import date

    from engine import run_model
    from export.api import build_bundle
    from export.workbook import WorkbookBuilder, registered_sheets as specs

    project, terms = ProjectInputs(), DebtTerms()
    model = build_bundle(
        project, terms, run_model(project, terms), generated_on=date(2026, 8, 6)
    )
    wb = WorkbookBuilder()
    for spec in sorted(specs(), key=lambda s: (s.build_order, s.name)):
        spec.builder(wb, model)
    return wb


@pytest.mark.parametrize("sheet,key", _MUST_BE_FORMULAS)
def test_model_rows_contain_formulas_not_numbers(sheet, key):
    path, _ = build_case("base")
    builder = _row_addresses(path)
    ref = builder.row(key)
    wb = load_workbook(path)
    ws = wb[sheet]
    for period in range(ref.start_period, ref.start_period + ref.n_periods):
        value = ws.cell(row=ref.row, column=ref.column(period)).value
        assert isinstance(value, str) and value.startswith("="), (
            f"{sheet}!{key} period {period} holds {value!r}, not a formula - "
            "the workbook would not recalculate"
        )


def test_the_debt_quantum_itself_is_a_formula():
    """Senior_Debt must fall out of the sizing tests, not be pasted in.

    This is the single most important cell in the workbook: if it were a
    number, changing the DSCR target in Excel would change nothing.
    """
    path, _ = build_case("base")
    wb = load_workbook(path)
    ref = wb.defined_names["Senior_Debt"].attr_text
    sheet, _, address = ref.partition("!")
    value = wb[sheet.strip("'")][address.replace("$", "")].value
    assert isinstance(value, str) and value.startswith("=MIN(")
    for test in (
        "Debt_DSCR_Test",
        "Debt_Gearing_Test",
        "Debt_LLCR_Test",
        "Debt_PLCR_Test",
    ):
        assert test in value


def test_only_one_solver_derived_cell_exists_and_it_is_flagged():
    """Everything except the applied grace period is re-derivable in Excel."""
    path, _ = build_case("base")
    wb = load_workbook(path)
    ws = wb["Inputs"]
    amber = [
        cell
        for row in ws.iter_rows()
        for cell in row
        if cell.fill is not None
        and cell.fill.fgColor is not None
        and cell.fill.fgColor.rgb == "FFFCE4D6"
    ]
    assert len(amber) == 1, [c.coordinate for c in amber]
    assert amber[0].coordinate == wb.defined_names["Grace_Periods"].attr_text.split(
        "!"
    )[1].replace("$", "")


# ---------------------------------------------------------------------------
# Named ranges
# ---------------------------------------------------------------------------

#: SPEC §6.5: "Named ranges for every driver." Every input a user would want to
#: change must be reachable by name.
_REQUIRED_DRIVER_NAMES = [
    "Capacity_MW",
    "Capex",
    "Construction_Months",
    "COD_Date",
    "Project_Life_Years",
    "Periods_Per_Year",
    "Production_Year1_MWh",
    "Degradation",
    "Contracted_Price",
    "Contracted_Share",
    "Contracted_Escalation",
    "Contract_Years",
    "Merchant_Price",
    "Merchant_Escalation",
    "Opex_Year1",
    "Opex_Escalation",
    "Tax_Rate",
    "Tax_Treatment_Code",
    "Depreciation_Years",
    "Target_DSCR",
    "Tenor_Years",
    "Interest_Rate",
    "Upfront_Fee_Pct",
    "Commitment_Fee_Pct",
    "Amort_Style_Code",
    "Grace_Period_Months",
    "Max_Gearing",
    "Tail_Years",
    "DSRA_Months",
    "Cash_Sweep_Pct",
    "Covenant_DSCR",
    "Lockup_DSCR",
    "Min_LLCR",
    "Min_PLCR",
    "Discount_Rate",
]

_REQUIRED_OUTPUT_NAMES = [
    "Senior_Debt",
    "Binding_Constraint",
    "Total_Project_Cost",
    "IDC",
    "Upfront_Fee_Amount",
    "Commitment_Fee_Total",
    "DSRA_Initial",
    "Debt_At_COD",
    "Equity_At_COD",
    "Gearing_Achieved",
    "Min_DSCR_Achieved",
    "LLCR_At_COD",
    "PLCR_At_COD",
    "Debt_WAL_Years",
    "Equity_IRR",
    "Equity_XIRR",
    "Equity_NPV",
    "Equity_MOIC",
    "Equity_Payback_Years",
    "Effective_Cost_Of_Debt",
    "After_Tax_Cost_Of_Debt",
    "WACC",
]


@pytest.mark.parametrize("name", _REQUIRED_DRIVER_NAMES + _REQUIRED_OUTPUT_NAMES)
def test_named_range_exists_and_points_at_a_real_cell(name):
    path, _ = build_case("base")
    wb = load_workbook(path)
    assert name in wb.defined_names, f"{name} is not a defined name"
    ref = wb.defined_names[name].attr_text
    sheet, _, address = ref.partition("!")
    sheet = sheet.strip("'")
    assert sheet in wb.sheetnames, ref
    assert re.fullmatch(r"\$[A-Z]{1,3}\$[0-9]+", address), ref
    assert wb[sheet][address.replace("$", "")].value is not None, (
        f"{name} points at {ref}, which is empty"
    )


def test_driver_names_point_at_blue_input_cells():
    """A driver that is secretly a formula is not a driver."""
    path, _ = build_case("base")
    wb = load_workbook(path)
    for name in _REQUIRED_DRIVER_NAMES:
        ref = wb.defined_names[name].attr_text
        sheet, _, address = ref.partition("!")
        cell = wb[sheet.strip("'")][address.replace("$", "")]
        assert cell.font.color is not None and cell.font.color.rgb == "FF0000CC", (
            f"{name} at {ref} is not styled as an input"
        )
        assert not (isinstance(cell.value, str) and cell.value.startswith("=")), (
            f"{name} at {ref} holds a formula, so it cannot be a driver"
        )


def test_series_named_ranges_span_the_whole_model_horizon():
    model, result = evaluated_case("base")
    n = result[0].cashflow.n_periods
    for name in (
        "CFADS_Series",
        "Revenue_Series",
        "EBITDA_Series",
        "Senior_Interest_Series",
        "Senior_Principal_Series",
        "Senior_Debt_Service_Series",
        "Debt_Closing_Balance_Series",
        "Distributions_Series",
    ):
        assert len(model.series(name)) == n, name
    assert len(model.series("Equity_Cashflow_Series")) == n + 1


# ---------------------------------------------------------------------------
# Presentation
# ---------------------------------------------------------------------------


def test_summary_sheets_hide_gridlines_and_freeze_panes():
    path, _ = build_case("base")
    wb = load_workbook(path)
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws.sheet_view.showGridLines is False, name
        assert ws.freeze_panes is not None, name


def test_column_widths_are_set_so_nothing_shows_as_hashes():
    path, _ = build_case("base")
    wb = load_workbook(path)
    for name in ("Operations", "Debt", "Waterfall", "Returns"):
        ws = wb[name]
        assert ws.column_dimensions["A"].width >= 40, name
        assert ws.column_dimensions["D"].width >= 10, name


def test_money_rows_use_the_banker_number_format():
    path, _ = build_case("base")
    wb = load_workbook(path)
    ws = wb["Debt"]
    formats = {
        cell.number_format
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "#,##0;(#,##0)" in formats
    assert '0.000"x"' in formats


@pytest.mark.parametrize(
    "case,project,terms",
    [
        (
            "level",
            ProjectInputs(),
            DebtTerms(amortization=AmortizationStyle.LEVEL),
        ),
        (
            "full_tax",
            ProjectInputs(tax_treatment=TaxTreatment.FULL),
            DebtTerms(),
        ),
    ],
)
def test_variant_workbooks_still_declare_iterative_calculation(case, project, terms):
    path, _ = build_case(case, project, terms)
    assert 'iterate="1"' in _calc_pr(path)
