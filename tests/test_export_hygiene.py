"""Hygiene: no answer keys, no personal data, and the disclaimer is present.

The hygiene rules: "not advice" appears as a prominent disclaimer; public
sources only; no real transaction's assumptions, no employer data and no PII
anywhere in the workbook.

The "answer key" test matters most: if an engine output were quietly pasted
into a cell, the workbook would agree with the engine for the wrong reason and
would stop agreeing the moment a user changed an input.
"""

from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from export.model import DISCLAIMER
from test_export_evaluator import build_case

_EMAIL = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
_UNIX_PATH = re.compile(r"/(?:Users|home|var|private)/")
_WINDOWS_PATH = re.compile(r"[A-Za-z]:\\")
_SSN = re.compile(r"\b\d{3}-\d{2}-\d{4}\b")
_LONG_DIGITS = re.compile(r"\b\d{9,}\b")


def _all_strings(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    yield ws.title, cell.coordinate, cell.value


def _static_numbers(path):
    """Every numeric cell that is not a formula - i.e. genuinely hard-coded."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, bool):
                    continue
                if isinstance(value, (int, float)):
                    yield ws.title, cell.coordinate, float(value)


# ---------------------------------------------------------------------------
# No answer keys
# ---------------------------------------------------------------------------


def test_no_engine_output_is_hard_coded_anywhere_in_the_workbook():
    """Every headline output must be computed, never pasted.

    A pasted answer would make the workbook agree with the engine on the day it
    was generated and disagree with itself the moment anyone touched an input.
    """
    path, (solution, waterfall, returns) = build_case("base")
    forbidden = {
        "senior debt": solution.debt_size,
        "total project cost": solution.construction.total_project_cost,
        "IDC": solution.construction.idc,
        "upfront fee": solution.construction.upfront_fee,
        "commitment fee": solution.construction.commitment_fee,
        "initial DSRA": solution.construction.dsra_initial,
        "equity at COD": solution.construction.equity_at_cod,
        "debt at COD": solution.construction.debt_at_cod,
        "minimum DSCR": solution.sizing.min_dscr,
        "LLCR": solution.sizing.llcr,
        "PLCR": solution.sizing.plcr,
        # Gearing is deliberately absent: on a gearing-bound deal the achieved
        # gearing equals the Max_Gearing *input* exactly, so scanning for it
        # would flag a legitimate input cell.
        "equity NPV": returns.equity_npv,
        "equity IRR": returns.equity_irr_post_tax,
        "equity MOIC": returns.equity_moic,
        "first period CFADS": solution.cashflow.cfads[0],
        "first debt service": solution.sizing.debt.debt_service[0],
        "first distribution": waterfall.distributions[0],
    }
    statics = list(_static_numbers(path))
    for label, value in forbidden.items():
        if value is None or abs(value) < 1.0:
            # Small dimensionless outputs (ratios, IRRs) are checked below with
            # a tight absolute tolerance instead of a relative one.
            matches = [
                (s, c, v)
                for s, c, v in statics
                if value is not None and abs(v - value) < 1e-9
            ]
        else:
            matches = [
                (s, c, v)
                for s, c, v in statics
                if abs(v - value) <= abs(value) * 1e-6
            ]
        assert not matches, (
            f"the {label} appears as a hard-coded number at {matches[:3]} - "
            "it must be a formula"
        )


def test_the_only_hard_coded_numbers_are_inputs_indices_and_weights():
    """Sanity-bound the amount of hard-coding, so it cannot creep.

    Static numbers legitimately exist for: the Inputs sheet, period and month
    index strips, the capex S-curve weights, and the zeroed MRA / subordinated
    input rows. Nothing else.
    """
    path, _ = build_case("base")
    wb = load_workbook(path)
    allowed_sheets = {"Inputs"}
    offenders = []
    for sheet, coordinate, _value in _static_numbers(path):
        if sheet in allowed_sheets:
            continue
        cell = wb[sheet][coordinate]
        fill = cell.fill
        if fill is not None and fill.fgColor is not None and fill.fgColor.rgb == "FF1F3864":
            continue  # the navy period-number band at the top of every sheet
        label = wb[sheet].cell(row=cell.row, column=1).value or ""
        if isinstance(label, str) and (
            label in ("Period", "Month")
            or label.startswith("Capex S-curve weight")
            or label.startswith("MRA ")
            or label.startswith("Subordinated debt service scheduled")
            or label.startswith("Target DSCR")
        ):
            continue
        offenders.append((sheet, coordinate, label))
    assert not offenders, offenders


# ---------------------------------------------------------------------------
# No personal or confidential data
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "pattern,description",
    [
        (_EMAIL, "an email address"),
        (_UNIX_PATH, "a local filesystem path"),
        (_WINDOWS_PATH, "a local filesystem path"),
        (_SSN, "something shaped like a national identifier"),
    ],
)
def test_no_personal_data_leaks_into_any_cell(pattern, description):
    path, _ = build_case("base")
    hits = [
        (sheet, coord, text)
        for sheet, coord, text in _all_strings(path)
        if pattern.search(text)
    ]
    assert not hits, f"{description} found: {hits[:3]}"


def test_no_account_length_digit_strings_in_any_cell():
    path, _ = build_case("base")
    hits = [
        (sheet, coord, text)
        for sheet, coord, text in _all_strings(path)
        if _LONG_DIGITS.search(text)
    ]
    assert not hits, hits[:3]


def test_workbook_metadata_carries_no_author_identity():
    path, _ = build_case("base")
    wb = load_workbook(path)
    for field in (wb.properties.creator, wb.properties.lastModifiedBy):
        assert field in (None, "", "openpyxl"), field


def test_no_sheet_mentions_a_real_counterparty_or_employer():
    """Every default must be a published market benchmark."""
    path, _ = build_case("base")
    text = " ".join(t for _, _, t in _all_strings(path)).lower()
    for token in ("envirospark", "internal use only", "apoorv", "singh"):
        assert token not in text, token


# ---------------------------------------------------------------------------
# Disclaimer and attribution
# ---------------------------------------------------------------------------


def test_the_disclaimer_appears_verbatim_on_both_summary_and_notes():
    path, _ = build_case("base")
    placements = {
        sheet for sheet, _, text in _all_strings(path) if DISCLAIMER in text
    }
    assert {"Summary", "Notes"} <= placements, placements


def test_the_disclaimer_says_the_four_things_it_has_to_say():
    for word in ("tax", "legal", "accounting", "investment"):
        assert word in DISCLAIMER.lower()
    assert "not" in DISCLAIMER.lower()
    assert "illustrative" in DISCLAIMER.lower()


def test_notes_sheet_documents_the_solver_derived_cell():
    path, _ = build_case("base")
    notes = " ".join(
        text for sheet, _, text in _all_strings(path) if sheet == "Notes"
    ).lower()
    assert "solver-derived" in notes
    assert "grace period" in notes
    assert "negative amortisation" in notes


def test_notes_sheet_declares_what_is_deferred_to_later_phases():
    path, _ = build_case("base")
    notes = " ".join(
        text for sheet, _, text in _all_strings(path) if sheet == "Notes"
    ).lower()
    for token in (
        "48e",
        "feoc",
        "material assistance cost ratio",
        "macrs",
        "partnership flip",
        "sale-leaseback",
        "6418",
        "704(b)",
    ):
        assert token in notes, token


def test_notes_sheet_attributes_its_market_defaults():
    path, _ = build_case("base")
    notes = " ".join(
        text for sheet, _, text in _all_strings(path) if sheet == "Notes"
    )
    assert "Norton Rose Fulbright" in notes
    assert "2026-01-29" in notes
    assert "IRC section 11(b)" in notes
    assert "MIT" in notes


def test_notes_sheet_states_the_methods_are_standard():
    """The workbook says plainly that the mathematics is standard practice."""
    path, _ = build_case("base")
    notes = " ".join(
        text for sheet, _, text in _all_strings(path) if sheet == "Notes"
    ).lower()
    assert "standard project finance" in notes


def test_notes_sheet_explains_iterative_calculation():
    path, _ = build_case("base")
    notes = " ".join(
        text for sheet, _, text in _all_strings(path) if sheet == "Notes"
    ).lower()
    assert "iterative calculation" in notes
    assert "circular" in notes
    assert "fullcalconload" in notes
