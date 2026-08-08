"""A minimal Excel formula evaluator, and its own tests.

WHY THIS EXISTS
---------------
The exported workbook carries *live
formulas* that reproduce the engine - not pasted numbers. Excel is not
available in this environment, so the only way to verify that claim honestly is
to read the emitted formulas back and evaluate them.

That is what this module does. It parses the subset of the Excel grammar the
exporter emits, resolves defined names and cross-sheet references, and then
calculates the workbook the way Excel does when iterative calculation is
switched on: repeatedly, in place, until nothing moves. The construction
funding circularity is resolved by exactly that mechanism, so the evaluator is
not merely checking arithmetic - it is checking that the circular chain
*converges*, and converges to the engine's answer.

It is deliberately small. It supports what the exporter writes and nothing
else; an unknown function raises rather than guessing.

Other export tests import from this module::

    from test_export_evaluator import evaluate_workbook
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Sequence

import pyxirr
import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

__all__ = ["EvaluatedWorkbook", "evaluate_workbook", "ExcelError"]


class ExcelError(Exception):
    """Raised for a formula this evaluator cannot handle."""


# ---------------------------------------------------------------------------
# Tokeniser
# ---------------------------------------------------------------------------

_CELL = r"\$?[A-Za-z]{1,3}\$?[0-9]+"
_TOKEN_RE = re.compile(
    rf"""
      (?P<ws>\s+)
    | (?P<string>"(?:[^"]|"")*")
    | (?P<quotedref>'(?:[^']|'')+'!{_CELL}(?::{_CELL})?)
    | (?P<sheetref>[A-Za-z_][A-Za-z0-9_.]*!{_CELL}(?::{_CELL})?)
    | (?P<localref>{_CELL}(?::{_CELL})?)
    | (?P<number>[0-9]+(?:\.[0-9]*)?(?:[eE][+-]?[0-9]+)?|\.[0-9]+)
    | (?P<name>[A-Za-z_][A-Za-z0-9_.]*)
    | (?P<op><=|>=|<>|[-+*/^(),<>=])
    """,
    re.VERBOSE,
)


@dataclass(frozen=True, slots=True)
class Token:
    kind: str
    text: str


def tokenize(formula: str) -> list[Token]:
    tokens: list[Token] = []
    pos = 0
    while pos < len(formula):
        match = _TOKEN_RE.match(formula, pos)
        if match is None:
            raise ExcelError(
                f"cannot tokenise at position {pos} of {formula!r}: "
                f"{formula[pos:pos + 20]!r}"
            )
        pos = match.end()
        kind = match.lastgroup
        assert kind is not None
        if kind == "ws":
            continue
        tokens.append(Token(kind, match.group()))
    return tokens


# ---------------------------------------------------------------------------
# Parser - recursive descent over the emitted subset of the Excel grammar
# ---------------------------------------------------------------------------

Node = tuple  # ("num", v) | ("str", s) | ("ref", sheet, a1) | ("name", n) |
#                ("call", fn, [args]) | ("bin", op, l, r) | ("neg", x)


class Parser:
    def __init__(self, tokens: list[Token], default_sheet: str) -> None:
        self.tokens = tokens
        self.pos = 0
        self.sheet = default_sheet

    def peek(self) -> Token | None:
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def take(self) -> Token:
        token = self.tokens[self.pos]
        self.pos += 1
        return token

    def expect(self, text: str) -> None:
        token = self.take()
        if token.text.upper() != text.upper():
            raise ExcelError(f"expected {text!r}, found {token.text!r}")

    def parse(self) -> Node:
        node = self.comparison()
        if self.pos != len(self.tokens):
            raise ExcelError(f"unconsumed tokens from {self.tokens[self.pos]!r}")
        return node

    def comparison(self) -> Node:
        node = self.additive()
        while (token := self.peek()) and token.text in ("=", "<>", "<", ">", "<=", ">="):
            self.take()
            node = ("bin", token.text, node, self.additive())
        return node

    def additive(self) -> Node:
        node = self.multiplicative()
        while (token := self.peek()) and token.text in ("+", "-"):
            self.take()
            node = ("bin", token.text, node, self.multiplicative())
        return node

    def multiplicative(self) -> Node:
        node = self.unary()
        while (token := self.peek()) and token.text in ("*", "/"):
            self.take()
            node = ("bin", token.text, node, self.unary())
        return node

    def unary(self) -> Node:
        token = self.peek()
        if token and token.text == "-":
            self.take()
            return ("neg", self.unary())
        if token and token.text == "+":
            self.take()
            return self.unary()
        return self.power()

    def power(self) -> Node:
        node = self.primary()
        token = self.peek()
        if token and token.text == "^":
            self.take()
            return ("bin", "^", node, self.unary())
        return node

    def primary(self) -> Node:
        token = self.take()
        if token.kind == "number":
            return ("num", float(token.text))
        if token.kind == "string":
            return ("str", token.text[1:-1].replace('""', '"'))
        if token.kind in ("quotedref", "sheetref"):
            sheet, _, a1 = token.text.partition("!")
            return ("ref", sheet.strip("'").replace("''", "'"), a1.replace("$", ""))
        if token.kind == "localref":
            return ("ref", self.sheet, token.text.replace("$", ""))
        if token.kind == "name":
            nxt = self.peek()
            if nxt and nxt.text == "(":
                self.take()
                args: list[Node] = []
                if (peeked := self.peek()) and peeked.text == ")":
                    self.take()
                    return ("call", token.text.upper(), args)
                while True:
                    args.append(self.comparison())
                    sep = self.take()
                    if sep.text == ")":
                        break
                    if sep.text != ",":
                        raise ExcelError(f"expected , or ) - found {sep.text!r}")
                return ("call", token.text.upper(), args)
            upper = token.text.upper()
            if upper in ("TRUE", "FALSE"):
                return ("num", 1.0 if upper == "TRUE" else 0.0)
            return ("name", token.text)
        if token.text == "(":
            node = self.comparison()
            self.expect(")")
            return node
        raise ExcelError(f"unexpected token {token.text!r}")


def parse_formula(formula: str, sheet: str) -> Node:
    body = formula[1:] if formula.startswith("=") else formula
    return Parser(tokenize(body), sheet).parse()


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

BLANK = None


def _numbers(values: Sequence[Any]) -> list[float]:
    """Numeric entries only. Excel ignores text and blanks inside a range."""
    out: list[float] = []
    for value in values:
        if isinstance(value, bool):
            continue
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            out.append(float(value))
    return out


def _as_number(value: Any) -> float:
    if value is BLANK or value == "":
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return float((value - datetime(1899, 12, 30)).days)
    if isinstance(value, list):
        return _as_number(value[0]) if value else 0.0
    raise ExcelError(f"cannot use {value!r} as a number")


def _flatten(value: Any) -> list[Any]:
    return value if isinstance(value, list) else [value]


def _edate(start: Any, months: float) -> datetime:
    if not isinstance(start, datetime):
        start = datetime(1899, 12, 30) + timedelta(days=_as_number(start))
    total = start.year * 12 + (start.month - 1) + int(round(months))
    year, month = divmod(total, 12)
    day = start.day
    while True:
        try:
            return datetime(year, month + 1, day)
        except ValueError:  # pragma: no cover - month-end rollback
            day -= 1


def _countif(values: Sequence[Any], criterion: Any) -> float:
    text = str(criterion)
    match = re.match(r"^(<=|>=|<>|<|>|=)?\s*(.*)$", text)
    assert match is not None
    op, operand = match.group(1) or "=", match.group(2)
    try:
        threshold = float(operand)
    except ValueError:
        return float(sum(1 for v in values if str(v) == operand))
    tests: dict[str, Callable[[float], bool]] = {
        "<": lambda x: x < threshold,
        ">": lambda x: x > threshold,
        "<=": lambda x: x <= threshold,
        ">=": lambda x: x >= threshold,
        "=": lambda x: x == threshold,
        "<>": lambda x: x != threshold,
    }
    return float(sum(1 for v in _numbers(values) if tests[op](v)))


def _npv(rate: float, values: Sequence[Any]) -> float:
    return sum(
        _as_number(v) / (1.0 + rate) ** (i + 1) for i, v in enumerate(values)
    )


def _irr(values: Sequence[Any]) -> float:
    flows = [_as_number(v) for v in values]
    if all(f >= 0 for f in flows) or all(f <= 0 for f in flows):
        return 0.0
    try:
        result = pyxirr.irr(flows)
    except Exception:  # pragma: no cover - degenerate series during iteration
        return 0.0
    return 0.0 if result is None else float(result)


def _xirr(values: Sequence[Any], dates: Sequence[Any]) -> float:
    flows = [_as_number(v) for v in values]
    if all(f >= 0 for f in flows) or all(f <= 0 for f in flows):
        return 0.0
    stamps: list[date] = []
    for d in dates:
        if isinstance(d, datetime):
            stamps.append(d.date())
        elif isinstance(d, date):
            stamps.append(d)
        else:
            stamps.append(
                (datetime(1899, 12, 30) + timedelta(days=_as_number(d))).date()
            )
    try:
        result = pyxirr.xirr(stamps, flows)
    except Exception:  # pragma: no cover
        return 0.0
    return 0.0 if result is None else float(result)


def _power(base: float, exponent: float) -> float:
    if base < 0 and exponent != int(exponent):
        return 0.0
    try:
        return float(base**exponent)
    except (OverflowError, ValueError, ZeroDivisionError):  # pragma: no cover
        return 0.0


# ---------------------------------------------------------------------------
# The workbook
# ---------------------------------------------------------------------------


@dataclass
class EvaluatedWorkbook:
    """A workbook read back from disk, with its formula graph calculated."""

    path: Path
    statics: dict[tuple[str, str], Any] = field(default_factory=dict)
    formulas: dict[tuple[str, str], Node] = field(default_factory=dict)
    raw_formulas: dict[tuple[str, str], str] = field(default_factory=dict)
    names: dict[str, str] = field(default_factory=dict)
    values: dict[tuple[str, str], Any] = field(default_factory=dict)
    order: list[tuple[str, str]] = field(default_factory=list)
    iterations: int = 0
    residual: float = math.inf

    # -- loading -------------------------------------------------------

    @classmethod
    def load(cls, path: str | Path) -> "EvaluatedWorkbook":
        wb = load_workbook(Path(path), data_only=False)
        model = cls(path=Path(path))
        for name, defined in wb.defined_names.items():
            model.names[name] = defined.attr_text
        for index, ws in enumerate(wb.worksheets):
            pending: list[tuple[int, int, int, tuple[str, str]]] = []
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    key = (ws.title, f"{cell.column_letter}{cell.row}")
                    if isinstance(value, str) and value.startswith("="):
                        model.formulas[key] = parse_formula(value, ws.title)
                        model.raw_formulas[key] = value
                        pending.append((index, cell.column, cell.row, key))
                        model.values[key] = 0.0
                    else:
                        model.statics[key] = value
                        model.values[key] = value
            # Column-major within a sheet, which is the natural calculation
            # order of a period-by-period financial model: every line item of
            # period t settles before period t+1 is touched. Excel builds a
            # real dependency chain; this is a cheap approximation of it and it
            # cuts the iteration count by roughly an order of magnitude.
            pending.sort()
            model.order.extend(key for *_, key in pending)
        return model

    # -- calculation ---------------------------------------------------

    def calculate(self, max_iterations: int = 300, tolerance: float = 1e-6) -> None:
        """Iterate the whole formula graph until nothing moves.

        This is what Excel does with ``iterate="1"``: recalculate in place,
        repeatedly, letting a circular chain converge. Gauss-Seidel (values are
        updated in place rather than double-buffered) converges faster and is
        what Excel actually does.
        """
        for iteration in range(1, max_iterations + 1):
            delta = 0.0
            for key in self.order:
                previous = self.values[key]
                current = self._evaluate(self.formulas[key], key[0])
                self.values[key] = current
                if isinstance(previous, (int, float)) and isinstance(
                    current, (int, float)
                ):
                    delta = max(delta, abs(float(current) - float(previous)))
                elif previous != current:
                    delta = math.inf
            self.iterations = iteration
            self.residual = delta
            if delta <= tolerance:
                return
        raise ExcelError(
            f"formula graph did not converge in {max_iterations} iterations "
            f"(largest change {self.residual:,.6g})"
        )

    # -- lookup --------------------------------------------------------

    def cell(self, sheet: str, address: str) -> Any:
        return self.values.get((sheet, address.replace("$", "")), BLANK)

    def name(self, name: str) -> Any:
        if name not in self.names:
            raise KeyError(f"no defined name {name!r} in {self.path.name}")
        return self._resolve(self.names[name])

    def series(self, name: str) -> list[float]:
        """A named horizontal range as a list of numbers."""
        return [_as_number(v) for v in _flatten(self.name(name))]

    # -- internals -----------------------------------------------------

    def _resolve(self, reference: str) -> Any:
        sheet, _, a1 = reference.partition("!")
        sheet = sheet.strip("'").replace("''", "'")
        return self._range(sheet, a1.replace("$", ""))

    def _range(self, sheet: str, a1: str) -> Any:
        if ":" not in a1:
            return self.values.get((sheet, a1), BLANK)
        start, end = a1.split(":")
        c0, r0 = _split(start)
        c1, r1 = _split(end)
        out: list[Any] = []
        for r in range(min(r0, r1), max(r0, r1) + 1):
            for c in range(min(c0, c1), max(c0, c1) + 1):
                out.append(
                    self.values.get((sheet, f"{get_column_letter(c)}{r}"), BLANK)
                )
        return out

    def _evaluate(self, node: Node, sheet: str) -> Any:
        kind = node[0]
        if kind == "num":
            return node[1]
        if kind == "str":
            return node[1]
        if kind == "ref":
            return self._range(node[1], node[2])
        if kind == "name":
            return self.name(node[1])
        if kind == "neg":
            return -_as_number(self._evaluate(node[1], sheet))
        if kind == "bin":
            return self._binary(node, sheet)
        if kind == "call":
            return self._call(node[1], node[2], sheet)
        raise ExcelError(f"unknown node {node!r}")  # pragma: no cover

    def _binary(self, node: Node, sheet: str) -> Any:
        _, op, left_node, right_node = node
        left = self._evaluate(left_node, sheet)
        right = self._evaluate(right_node, sheet)
        if op in ("=", "<>"):
            same = _loose_equal(left, right)
            return 1.0 if (same if op == "=" else not same) else 0.0
        a, b = _as_number(left), _as_number(right)
        if op == "+":
            return a + b
        if op == "-":
            return a - b
        if op == "*":
            return a * b
        if op == "/":
            # Excel raises #DIV/0!. During iteration a divisor is routinely zero
            # on the first pass, so a zero keeps the fixed point moving instead
            # of poisoning the whole graph.
            return a / b if b != 0 else 0.0
        if op == "^":
            return _power(a, b)
        comparisons = {
            "<": a < b, ">": a > b, "<=": a <= b, ">=": a >= b,
        }
        return 1.0 if comparisons[op] else 0.0

    def _call(self, fn: str, args: list[Node], sheet: str) -> Any:
        # Lazy functions: only the taken branch is evaluated, exactly as Excel.
        if fn == "IF":
            condition = _as_number(self._evaluate(args[0], sheet))
            if condition:
                return self._evaluate(args[1], sheet)
            if len(args) > 2:
                return self._evaluate(args[2], sheet)
            return 0.0
        if fn == "CHOOSE":
            index = int(round(_as_number(self._evaluate(args[0], sheet))))
            if not 1 <= index <= len(args) - 1:
                raise ExcelError(f"CHOOSE index {index} out of range")
            return self._evaluate(args[index], sheet)
        if fn == "IFERROR":
            try:
                return self._evaluate(args[0], sheet)
            except Exception:  # pragma: no cover
                return self._evaluate(args[1], sheet)

        values = [self._evaluate(arg, sheet) for arg in args]
        flat = [item for value in values for item in _flatten(value)]

        if fn == "SUM":
            return sum(_numbers(flat))
        if fn == "MIN":
            numbers = _numbers(flat)
            return min(numbers) if numbers else 0.0
        if fn == "MAX":
            numbers = _numbers(flat)
            return max(numbers) if numbers else 0.0
        if fn == "COUNT":
            return float(len(_numbers(flat)))
        if fn == "AND":
            return 1.0 if all(_as_number(v) for v in flat) else 0.0
        if fn == "OR":
            return 1.0 if any(_as_number(v) for v in flat) else 0.0
        if fn == "NOT":
            return 0.0 if _as_number(flat[0]) else 1.0
        if fn == "ABS":
            return abs(_as_number(values[0]))
        if fn == "INT":
            return float(math.floor(_as_number(values[0])))
        if fn == "ROUND":
            digits = int(_as_number(values[1]))
            return float(_round_half_up(_as_number(values[0]), digits))
        if fn == "SUMPRODUCT":
            columns = [_flatten(v) for v in values]
            total = 0.0
            for row in zip(*columns):
                product = 1.0
                for item in row:
                    product *= _as_number(item)
                total += product
            return total
        if fn == "NPV":
            return _npv(_as_number(values[0]), _flatten(values[1]))
        if fn == "IRR":
            return _irr(_flatten(values[0]))
        if fn == "XIRR":
            return _xirr(_flatten(values[0]), _flatten(values[1]))
        if fn == "INDEX":
            items = _flatten(values[0])
            index = int(round(_as_number(values[1])))
            if not 1 <= index <= len(items):
                return 0.0
            return items[index - 1]
        if fn == "COUNTIF":
            return _countif(_flatten(values[0]), values[1])
        if fn == "EDATE":
            return _edate(values[0], _as_number(values[1]))
        raise ExcelError(f"unsupported function {fn}()")


def _round_half_up(value: float, digits: int) -> float:
    scale = 10.0**digits
    return math.floor(abs(value) * scale + 0.5) / scale * (1 if value >= 0 else -1)


def _loose_equal(left: Any, right: Any) -> bool:
    if isinstance(left, str) or isinstance(right, str):
        if isinstance(left, str) and isinstance(right, str):
            return left.upper() == right.upper()
        return False
    return _as_number(left) == _as_number(right)


def _split(address: str) -> tuple[int, int]:
    match = re.match(r"^([A-Za-z]+)([0-9]+)$", address)
    if match is None:  # pragma: no cover
        raise ExcelError(f"bad cell address {address!r}")
    return column_index_from_string(match.group(1)), int(match.group(2))


def evaluate_workbook(path: str | Path, **kwargs) -> EvaluatedWorkbook:
    """Load a workbook and calculate it. The one entry point other tests use."""
    model = EvaluatedWorkbook.load(path)
    model.calculate(**kwargs)
    return model


# ===========================================================================
# Tests for the evaluator itself
# ===========================================================================


def _eval(formula: str, cells: dict[str, Any] | None = None) -> Any:
    model = EvaluatedWorkbook(path=Path("memory"))
    for address, value in (cells or {}).items():
        model.values[("S", address)] = value
    return model._evaluate(parse_formula(formula, "S"), "S")


@pytest.mark.parametrize(
    "formula,expected",
    [
        ("=1+2*3", 7.0),
        ("=(1+2)*3", 9.0),
        ("=2^10", 1024.0),
        ("=(1+0.05)^-2", 1.0 / 1.05**2),
        ("=-3+1", -2.0),
        ("=1E+30", 1e30),
        ("=MIN(3,1,2)", 1.0),
        ("=MAX(3,1,2)", 3.0),
        ("=IF(1>2,10,20)", 20.0),
        ("=IF(2>1,10,20)", 10.0),
        ("=AND(1>0,2>1)", 1.0),
        ("=OR(1>2,2>3)", 0.0),
        ("=ROUND(2.5,0)", 3.0),
        ("=ROUND(18.4,0)", 18.0),
        ("=INT(7/2)", 3.0),
        ("=ABS(-4)", 4.0),
        ("=CHOOSE(2,10,20,30)", 20.0),
    ],
)
def test_evaluator_arithmetic(formula, expected):
    assert _eval(formula) == pytest.approx(expected)


def test_evaluator_ranges_and_functions():
    cells = {"A1": 1.0, "B1": 2.0, "C1": 3.0, "A2": 10.0, "B2": 20.0, "C2": 30.0}
    assert _eval("=SUM(A1:C1)", cells) == pytest.approx(6.0)
    assert _eval("=SUMPRODUCT(A1:C1,A2:C2)", cells) == pytest.approx(140.0)
    assert _eval("=INDEX(A2:C2,3)", cells) == pytest.approx(30.0)
    assert _eval("=MIN(A1:C1)", cells) == pytest.approx(1.0)


def test_evaluator_ignores_text_in_ranges_like_excel():
    cells = {"A1": 5.0, "B1": "", "C1": 2.0}
    assert _eval("=MIN(A1:C1)", cells) == pytest.approx(2.0)
    assert _eval("=SUM(A1:C1)", cells) == pytest.approx(7.0)


def test_evaluator_npv_matches_the_engine_convention():
    """NPV discounts its first argument by one period - the same convention
    ``engine.debt.present_value`` uses with a zero offset."""
    from engine.debt import present_value

    cells = {"A1": 100.0, "B1": 110.0, "C1": 120.0}
    got = _eval("=NPV(0.07,A1:C1)", cells)
    assert got == pytest.approx(present_value((100.0, 110.0, 120.0), 0.07))


def test_evaluator_irr_matches_the_engine():
    from engine.metrics import irr

    cells = {"A1": -1000.0, "B1": 400.0, "C1": 400.0, "D1": 400.0}
    assert _eval("=IRR(A1:D1)", cells) == pytest.approx(irr((-1000.0, 400.0, 400.0, 400.0)))


def test_evaluator_countif_supports_the_payback_idiom():
    cells = {"A1": -5.0, "B1": -2.0, "C1": 3.0}
    assert _eval('=COUNTIF(A1:C1,"<0")', cells) == pytest.approx(2.0)


def test_evaluator_edate_advances_whole_months():
    cells = {"A1": datetime(2027, 1, 1)}
    assert _eval("=EDATE(A1,12)", cells) == datetime(2028, 1, 1)
    assert _eval("=EDATE(A1,6)", cells) == datetime(2027, 7, 1)


def test_evaluator_resolves_a_circular_pair():
    """The evaluator must behave like Excel with iterate=1: a circular pair
    converges rather than raising."""
    model = EvaluatedWorkbook(path=Path("memory"))
    for address, formula in (("A1", "=10+0.5*B1"), ("B1", "=0.5*A1")):
        model.formulas[("S", address)] = parse_formula(formula, "S")
        model.values[("S", address)] = 0.0
        model.order.append(("S", address))
    model.calculate()
    # A = 10 + 0.5B, B = 0.5A  ->  A = 40/3, B = 20/3
    assert model.cell("S", "A1") == pytest.approx(40.0 / 3.0, abs=1e-5)
    assert model.cell("S", "B1") == pytest.approx(20.0 / 3.0, abs=1e-5)


def test_evaluator_rejects_an_unknown_function():
    with pytest.raises(ExcelError, match="unsupported function"):
        _eval("=VLOOKUP(1,A1:B2,2,0)", {})


# ===========================================================================
# Shared build helpers, used by the other export test modules
# ===========================================================================

import tempfile  # noqa: E402  (kept next to the helpers that use it)

from engine import DebtTerms, ProjectInputs, run_model  # noqa: E402
from export import build_workbook  # noqa: E402

#: Workbooks are built into a throwaway directory, never into the repository.
_WORKBOOK_DIR = Path(tempfile.mkdtemp(prefix="structura-export-"))
_BUILT: dict[str, tuple[Path, tuple]] = {}
_EVALUATED: dict[str, "EvaluatedWorkbook"] = {}

#: Fixed so two runs produce byte-identical files.
GENERATED_ON = date(2026, 8, 6)


def build_case(
    name: str,
    project: ProjectInputs | None = None,
    terms: DebtTerms | None = None,
    **kwargs,
) -> tuple[Path, tuple]:
    """Build the workbook for a named case once and cache it.

    Returns the file path and the engine result it was built from, so a test
    can compare the two without re-running either.
    """
    if name not in _BUILT:
        project = project or ProjectInputs()
        terms = terms or DebtTerms()
        result = run_model(
            project,
            terms,
            lockup_dscr=kwargs.get("lockup_dscr"),
            discount_rate=kwargs.get("discount_rate", 0.10),
        )
        path = build_workbook(
            project,
            terms,
            result,
            _WORKBOOK_DIR / f"{name}.xlsx",
            generated_on=GENERATED_ON,
            **kwargs,
        )
        _BUILT[name] = (path, result)
    return _BUILT[name]


def evaluated_case(
    name: str,
    project: ProjectInputs | None = None,
    terms: DebtTerms | None = None,
    **kwargs,
) -> tuple["EvaluatedWorkbook", tuple]:
    """``build_case`` plus a calculated formula graph, cached."""
    path, result = build_case(name, project, terms, **kwargs)
    if name not in _EVALUATED:
        _EVALUATED[name] = evaluate_workbook(path)
    return _EVALUATED[name], result


def test_build_helper_writes_outside_the_repository():
    path, _ = build_case("base")
    assert path.exists() and path.suffix == ".xlsx"
    assert "structura/app/export" not in str(path)
    assert "structura/app/tests" not in str(path)
